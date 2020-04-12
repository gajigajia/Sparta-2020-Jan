import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

#data_only=true로 해야 수식이 아닌 값으로 받아옴
load_wb = load_workbook("mymusic.xlsx",data_only=True)
load_ws = load_wb['Sheet1']

headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
data = requests.get('https://www.genie.co.kr/chart/top200?ditc=D&ymd=20200403&hh=23&rtm=N&pg=1',headers=headers)

soup = BeautifulSoup(data.text, 'html.parser')

songs = soup.select('#body-content > div.newest-list > div > table > tbody > tr')
#body-content > div.newest-list > div > table > tbody > tr:nth-child(1) > td.number

i=2
for song in songs:
    rank   = song.select_one('td.number').text[0:2].strip()
    title  = song.select_one('td.info > a.title.ellipsis').text.strip()
    artist = song.select_one('td.info > a.artist.ellipsis').text

    print(rank,title,artist)

    load_ws.cell(i,1,rank)
    load_ws.cell(i,2,title)
    load_ws.cell(i,3,artist)

    i += 1

load_wb.save("mymusic.xlsx")

